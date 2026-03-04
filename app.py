import streamlit as st
import pandas as pd
import sqlite3
import re
import unicodedata
from rapidfuzz import process, fuzz
import xml.etree.ElementTree as ET
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- CONFIGURAÇÃO VISUAL DO APP ---
st.set_page_config(page_title="FLV Enterprise - Tome Leve", page_icon="🍎", layout="wide")

st.markdown("""
    <style>
    div.stButton > button:first-child {
        background-color: #002060; color: white; height: 3em; font-weight: bold; width: 100%; border-radius: 8px;
    }
    div.stButton > button:first-child:hover { background-color: #00133d; }
    </style>
""", unsafe_allow_html=True)

st.title("🍎 Sistema Integrado FLV Enterprise")

NAMESPACE_NFE = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
DB_NAME = "auditoria_flv_app_v25.db" 
TOLERANCIA_DIF = 0.001

def normalizar(texto):
    if pd.isna(texto) or texto is None: return ""
    texto = str(texto).upper().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    texto = re.sub(r'[^\w\s\.\-]', '', texto)
    if "MACA GALA GRANEL P" in texto: return "MACA BABY KG"
    if "MACA GALA PREMIUM" in texto or "TP 135" in texto: return "MACA GALA KG"
    return texto

def descobrir_loja(cnpj_dest, nome_dest):
    nome = normalizar(nome_dest)
    cnpj = ''.join(filter(str.isdigit, str(cnpj_dest)))
    if 'LOJA 01' in nome or 'LOJA-1' in nome or '( 01 )' in nome: return 'Loja_1'
    if 'LOJA 02' in nome or 'LOJA-2' in nome or '( 02 )' in nome: return 'Loja_2'
    if 'LOJA 03' in nome or 'LOJA-3' in nome or '( 03 )' in nome: return 'Loja_3'
    if 'LOJA 05' in nome or 'LOJA-5' in nome or '( 05 )' in nome: return 'Loja_5'
    if cnpj.endswith('000100'): return 'Loja_1'
    if cnpj.endswith('000363'): return 'Loja_2'
    if cnpj.endswith('000444'): return 'Loja_3'
    if cnpj.endswith('000606'): return 'Loja_5'
    if cnpj.endswith('000101'): return 'Loja_6'
    if cnpj.endswith('000365'): return 'Loja_7'
    if 'BARRETOS' in nome: return 'Loja_6'
    if 'COLINA' in nome or 'ANGELICOLA' in nome or cnpj.endswith('000184'): return 'Loja_8'
    return 'Loja_Desconhecida'

def traduzir_fornecedor(nome_bruto):
    nome = normalizar(nome_bruto)
    if 'RASTEIRA' in nome or 'RIBER' in nome: return 'RIBER FRUTAS'
    if 'HERCULES' in nome or 'RICARDO' in nome: return 'RICARDO'
    if 'CLAUDIO MARCELO' in nome or 'MARCELO' in nome: return 'MARCELO MILHO'
    if '2A COMERCIO' in nome or 'PIMENTA' in nome or '2 A COMERCIO' in nome: return 'IRMAOS PIMENTA'
    if 'ND COMERCIO' in nome or ' ND ' in f" {nome} " or nome == 'ND' or 'N D COM' in nome or 'N.D' in nome: return 'ND'
    if 'NICOLETI' in nome: return 'NICOLETI'
    if 'COAL' in nome or 'ARANDA' in nome: return 'COAL'
    if 'DRUB' in nome or 'ADILSON' in nome: return 'DRUB'
    if 'ZERO' in nome.split() or 'FRUTAS ZERO' in nome: return 'FRUTAS ZERO'
    if 'TAIS' in nome.split(): return 'TAIS'
    if 'LUCIO' in nome: return 'LUCIO ORLANDO'
    return nome.replace("FORNECEDOR", "").strip()

def descobrir_familia(nome):
    n = normalizar(nome)
    if "MELANCIA" in n and ("BABY" in n or "MINI" in n): return "MELANCIA_BABY"
    if "MELANCIA" in n: return "MELANCIA"
    if "BANANA" in n and "NANICA" in n: return "BANANA_NANICA"
    if "BANANA" in n and "PRATA" in n: return "BANANA_PRATA"
    if "BANANA" in n and "MACA" in n: return "BANANA_MACA"
    if "BANANA" in n and "MARMELO" in n: return "BANANA_MARMELO"
    if "BANANA" in n and "DA TERRA" in n: return "BANANA_TERRA"
    if "BANANA" in n: return "BANANA_OUTRA"
    if "BATATA" in n and "DOCE" in n: return "BATATA_DOCE"
    if "BATATA" in n: return "BATATA"
    if "CEBOLA" in n and "ROXA" in n: return "CEBOLA_ROXA"
    if "CEBOLA" in n: return "CEBOLA"
    if "ALHO" in n and "ROXO" in n: return "ALHO_ROXO"
    if "ALHO" in n: return "ALHO"
    if "PIMENTAO" in n and "VERMELHO" in n: return "PIMENTAO_VERMELHO"
    if "PIMENTAO" in n and "AMARELO" in n: return "PIMENTAO_AMARELO"
    if "PIMENTAO" in n and "VERDE" in n: return "PIMENTAO_VERDE"
    if "COLORIDO" in n or "COLCORIDO" in n: return "PIMENTAO_COLORIDO"
    if "PIMENTAO" in n: return "PIMENTAO_OUTRO"
    if "PAPAIA" in n or "PAPAYA" in n: return "MAMAO_PAPAIA"
    if "FORMOSA" in n: return "MAMAO_FORMOSA"
    if "MAMAO" in n: return "MAMAO_OUTRO"
    if "MORANGO" in n or "MORANGUINHO" in n: return "MORANGO"
    regras_simples = [
        "MELAO", "LARANJA", "LIMAO", "TANGERINA", "PONKAN", "MURCOTE",
        "CHUCHU", "CENOURA", "BETERRABA", "BERINJELA", "REPOLHO", "COUVE FLOR", "COUVE",
        "BROCOLIS", "TOMATE", "MACA", "PERA", "MANGA", "ABACATE", "ABACAXI", "QUIABO",
        "PEPINO", "MARACUJA", "MILHO", "VAGEM", "JILO", "KIWI", "GENGIBRE", "GOIABA",
        "INHAME", "SALSAO", "RABANETE", "AIPIM"
    ]
    for r in regras_simples:
        if r in n:
            if r in ["PONKAN", "MURCOTE"]: return "TANGERINA"
            return r
    if "ABOBRINHA" in n or "ABOBORA" in n or "CABOTIA" in n: return "ABOBORA_ABOBRINHA"
    return n.split()[0] if n else ""

def criar_banco():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria_v2 (
            id INTEGER PRIMARY KEY AUTOINCREMENT, id_execucao TEXT, loja TEXT, fornecedor TEXT,
            produto_pedido TEXT, produto_xml TEXT, qtd_pedido REAL, qtd_nota REAL,
            diferenca REAL, status_visual TEXT, status_codigo INTEGER,
            qtd_fisico TEXT, padrao_fisico TEXT, status_doca TEXT, diferenca_doca REAL
        )
        """)
        conn.commit()

def classificar(qtd_ped, qtd_fat, tipo):
    if tipo == "SEM_FORNECEDOR": return ("⚪ SEM NFe P/ FORN", 98, -qtd_ped)
    if tipo == "SEM_PRODUTO": return ("⚪ PRODUTO NÃO FATURADO", 99, -qtd_ped)
    diferenca = qtd_fat - qtd_ped
    if abs(diferenca) < TOLERANCIA_DIF: return ("🟢 OK", 0, 0.0)
    if diferenca < 0: return (f"🔴 NFe FALTA {abs(diferenca):.2f}".replace('.00',''), -1, diferenca)
    return (f"🟡 NFe SOBRA {diferenca:.2f}".replace('.00',''), 1, diferenca)

def classificar_doca(qtd_xml, qtd_fisico):
    diferenca = qtd_fisico - qtd_xml
    if abs(diferenca) < TOLERANCIA_DIF: return "🟢 FÍSICO BATEU"
    if diferenca < 0: return f"🔴 FÍSICO FALTOU {abs(diferenca):.2f}".replace('.00','')
    return f"🟡 FÍSICO SOBROU {diferenca:.2f}".replace('.00','')

def gerar_excel_auditoria(df_final):
    df_final = df_final.fillna("")
    wb = Workbook()
    wb.remove(wb.active)
    for loja in sorted(df_final['loja'].unique()):
        df_loja = df_final[df_final['loja'] == loja].copy()
        ws = wb.create_sheet(title=loja)
        ws.append([f"AUDITORIA 3 VIAS - {loja.upper().replace('_', ' ')}"])
        ws.merge_cells('A1:H1')
        ws['A1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.append(['Produto Pedido', 'Qtd Ped', 'Produto NFe', 'Qtd NFe', 'Status NFe', 'Qtd Doca', 'Padrão', 'Status Doca (Físico x NFe)'])
        for cell in ws[2]: cell.font = Font(bold=True)
        current_forn = None
        for _, row in df_loja.iterrows():
            if row['fornecedor'] != current_forn:
                if current_forn is not None: ws.append([])
                current_forn = row['fornecedor']
                ws.append([f"Fornecedor: {current_forn}", "", "", "", "", "", "", ""])
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=8)
                for cell in ws[ws.max_row]:
                    if "FATURADO SEM PEDIDO" in current_forn:
                        cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                        cell.font = Font(color="60497A", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        cell.font = Font(color="002060", bold=True)
            ws.append([row['produto_pedido'], row['qtd_pedido'], row['produto_xml'], row['qtd_nota'], row['status_visual'], row['qtd_fisico'], row['padrao_fisico'], row['status_doca']])
            status_nfe_cell = ws.cell(row=ws.max_row, column=5)
            val_nfe = status_nfe_cell.value
            if val_nfe:
                if "🟢" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "🔴" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif "🟡" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "🔵" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
                elif "⚪" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                elif "🟣" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
            status_doca_cell = ws.cell(row=ws.max_row, column=8)
            val_doca = status_doca_cell.value
            if val_doca:
                if "🟢" in val_doca: status_doca_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "🔴" in val_doca: status_doca_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif "🟡" in val_doca: status_doca_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "⚪" in val_doca: status_doca_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 28
    return wb

aba_preparador, aba_auditoria = st.tabs(["🧹 1. Preparador de Pedidos", "🍎 2. Auditoria 3 Vias (NFe x Doca)"])

with aba_preparador:
    st.header("🧹 Preparador de Planilha do Comprador")
    arquivo_bruto = st.file_uploader("Arraste a Planilha Bruta (CSV ou Excel)", type=['csv', 'xlsx'], key="uploader_bruto")
    if st.button("Limpar e Preparar Planilha"):
        if arquivo_bruto:
            with st.spinner("Lendo estrutura e blindando..."):
                try:
                    nome_planilha = arquivo_bruto.name
                    if nome_planilha.endswith('.csv'): df_bruto = pd.read_csv(arquivo_bruto, header=None)
                    else:
                        todas_as_abas = pd.read_excel(arquivo_bruto, sheet_name=None, header=None)
                        abas_validas = [df for n, df in todas_as_abas.items() if str(n).lower() not in ['ped', 'com', 'sis'] and not str(n).isdigit()]
                        df_bruto = pd.concat(abas_validas, ignore_index=True) if abas_validas else pd.read_excel(arquivo_bruto, header=None)

                    lojas_alvo, coluna_padrao, coluna_custo = {}, -1, -1
                    for index, row in df_bruto.head(50).iterrows():
                        for col_idx, val in enumerate(row):
                            texto = str(val).strip().upper()
                            if texto == 'L1': lojas_alvo['Loja_1'] = col_idx
                            elif texto == 'L2': lojas_alvo['Loja_2'] = col_idx
                            elif texto == 'L3': lojas_alvo['Loja_3'] = col_idx
                            elif texto == 'L5': lojas_alvo['Loja_5'] = col_idx 
                            elif texto == 'L6': lojas_alvo['Loja_6'] = col_idx
                            elif texto == 'L7': lojas_alvo['Loja_7'] = col_idx
                            elif texto == 'L8': lojas_alvo['Loja_8'] = col_idx
                            elif 'PADRÃO' in texto or 'PADRAO' in texto: coluna_padrao = col_idx
                            elif 'CUSTO' in texto: coluna_custo = col_idx
                        if lojas_alvo: break

                    max_col = df_bruto.shape[1]
                    if coluna_padrao == -1: coluna_padrao = 10 if max_col > 10 else (max_col - 1)
                    if coluna_custo == -1: coluna_custo = 11 if max_col > 11 else (max_col - 1)

                    fornecedor_atual, cod_fornecedor_atual = "DESCONHECIDO", "-"
                    lista_fornecedores, lista_codigos = [], []
                    for index, row in df_bruto.iterrows():
                        col0_str, col1_str = str(row[0]).strip().upper(), str(row[1]).strip()
                        if "PEDIDO FLV" in col0_str:
                            nome_sujo = col0_str.replace("PEDIDO FLV", "").split("202")[0].replace(",", "").strip()
                            fornecedor_atual = nome_sujo if nome_sujo else "FORNECEDOR"
                        elif "CÓD" in col0_str and "FORN" in col0_str: cod_fornecedor_atual = col1_str
                        lista_fornecedores.append(fornecedor_atual)
                        lista_codigos.append(cod_fornecedor_atual)

                    df_bruto['Fornecedor'] = lista_fornecedores
                    df_bruto['Cod_Fornecedor'] = lista_codigos
                    df_bruto[0] = pd.to_numeric(df_bruto[0], errors='coerce')
                    df_dados = df_bruto.dropna(subset=[0]).copy()
                    df_dados[0] = df_dados[0].astype(int)

                    wb = Workbook()
                    wb.remove(wb.active)
                    fill_loja = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    font_loja = Font(color="FFFFFF", bold=True, size=14)
                    fill_fornecedor = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    font_fornecedor = Font(color="002060", bold=True, size=11)

                    for nome_loja, indice_coluna in lojas_alvo.items():
                        if indice_coluna >= max_col: continue
                        df_loja = df_dados[[0, 1, 'Cod_Fornecedor', 'Fornecedor', indice_coluna, coluna_padrao, coluna_custo]].copy()
                        df_loja.columns = ['Código', 'Descrição', 'Cod_Fornecedor', 'Fornecedor', 'Qtd_Pedida', 'Padrão_Cx', 'Custo']
                        df_loja['Qtd_Pedida'] = pd.to_numeric(df_loja['Qtd_Pedida'], errors='coerce')
                        df_loja['Custo'] = pd.to_numeric(df_loja['Custo'], errors='coerce').fillna(0)
                        df_loja = df_loja[df_loja['Qtd_Pedida'] > 0]
                        if df_loja.empty: continue
                        df_loja = df_loja.sort_values(by=['Fornecedor', 'Descrição'])
                        ws = wb.create_sheet(title=nome_loja)
                        ws.append([f"CONFERÊNCIA - {nome_loja.upper().replace('_', ' ')}"])
                        ws.merge_cells('A1:E1')
                        ws['A1'].fill = fill_loja
                        ws['A1'].font = font_loja
                        ws.append(['Código', 'Descrição', 'Qtd_Pedida', 'Padrão_Cx', 'Custo'])
                        for cell in ws[2]: cell.font = Font(bold=True)
                        current_forn = None
                        for _, row in df_loja.iterrows():
                            if row['Fornecedor'] != current_forn:
                                if current_forn is not None: ws.append([])
                                current_forn = row['Fornecedor']
                                cod_forn = row['Cod_Fornecedor']
                                ws.append([f"Fornecedor: {cod_forn} - {current_forn}", "", "", "", ""])
                                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
                                for cell in ws[ws.max_row]: cell.fill = fill_fornecedor; cell.font = font_fornecedor
                            ws.append([row['Código'], row['Descrição'], row['Qtd_Pedida'], row['Padrão_Cx'], row['Custo']])
                            ws.cell(row=ws.max_row, column=5).number_format = 'R$ #,##0.00'
                        ws.column_dimensions['B'].width = 45
                        for c in ['A','C','D','E']: ws.column_dimensions[c].width = 15
                    out_excel = io.BytesIO()
                    wb.save(out_excel)
                    st.success("✨ Planilha preparada!")
                    st.download_button(label="📥 Baixar Planilha", data=out_excel.getvalue(), file_name=f"Pedidos_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e: st.error(f"Erro: {e}")

with aba_auditoria:
    st.header("🍎 Cruzamento Triplo: Pedido vs XML vs Físico")
    col1, col2, col3 = st.columns(3)
    with col1: arquivo_excel = st.file_uploader("Arraste o Pedido", type=['xlsx'], key="uploader_pedidos")
    with col2: arquivos_xml = st.file_uploader("Arraste os XMLs", type=['xml'], accept_multiple_files=True, key="uploader_xmls")
    with col3: arquivos_contagem = st.file_uploader("Arraste o Romaneio", type=['xlsx', 'csv'], accept_multiple_files=True, key="uploader_contagem")

    if st.button("Executar Auditoria Tripla Implacável"):
        if not arquivo_excel or not arquivos_xml: st.warning("⚠️ Você precisa de pelo menos o Pedido e os XMLs.")
        else:
            with st.spinner("Processando os fluxos de dados..."):
                criar_banco()
                id_execucao = datetime.now().strftime("%Y%m%d%H%M%S")
                try:
                    df_pedidos_raw = pd.read_excel(arquivo_excel, sheet_name=None, header=None)
                    pedidos_lista = []
                    for aba, df in df_pedidos_raw.items():
                        forn_orig, forn_macro = "DESCONHECIDO", "DESCONHECIDO"
                        for _, row in df.iterrows():
                            col0 = str(row[0]).strip()
                            if col0.startswith("Fornecedor:"):
                                forn_orig = col0.replace("Fornecedor:", "").strip()
                                forn_macro = traduzir_fornecedor(forn_orig)
                            else:
                                val = pd.to_numeric(col0, errors='coerce')
                                if pd.notna(val) and val > 0:
                                    pedidos_lista.append({'Loja': aba, 'Fornecedor_Original': forn_orig, 'Fornecedor_Macro': forn_macro, 'Produto': normalizar(row[1]), 'Qtd': float(row[2]) if pd.notna(row[2]) else 0.0})
                    df_pedidos = pd.DataFrame(pedidos_lista).groupby(['Loja', 'Fornecedor_Original', 'Fornecedor_Macro', 'Produto'], as_index=False)['Qtd'].sum()

                    notas = []
                    for xml_file in arquivos_xml:
                        try:
                            tree = ET.parse(io.BytesIO(xml_file.read()))
                            inf = tree.getroot().find('.//nfe:infNFe', NAMESPACE_NFE)
                            if inf is None: continue
                            emit_node = inf.find('nfe:emit/nfe:xNome', NAMESPACE_NFE)
                            dest_node = inf.find('nfe:dest/nfe:CNPJ', NAMESPACE_NFE)
                            dest_nome_node = inf.find('nfe:dest/nfe:xNome', NAMESPACE_NFE)
                            forn_macro = traduzir_fornecedor(emit_node.text) if emit_node is not None else "DESCONHECIDO"
                            loja_xml = descobrir_loja(dest_node.text if dest_node is not None else "0", dest_nome_node.text if dest_nome_node is not None else "")
                            for det in inf.findall('nfe:det', NAMESPACE_NFE):
                                prod_node = det.find('nfe:prod/nfe:xProd', NAMESPACE_NFE)
                                qtd_node = det.find('nfe:prod/nfe:qCom', NAMESPACE_NFE)
                                if prod_node is not None and qtd_node is not None:
                                    notas.append({"Loja": loja_xml, "Fornecedor_Macro": forn_macro, "Produto": normalizar(prod_node.text), "Qtd": float(qtd_node.text)})
                        except: pass
                    df_notas_agg = pd.DataFrame(notas).groupby(['Loja', 'Fornecedor_Macro', 'Produto'], as_index=False)['Qtd'].sum() if notas else pd.DataFrame()

                    contagens_lista = []
                    if arquivos_contagem:
                        for f in arquivos_contagem:
                            try:
                                df_c = pd.DataFrame()
                                if f.name.lower().endswith('.csv'):
                                    for sep, enc in [(',', 'utf-8'), (';', 'utf-8'), (',', 'latin1'), (';', 'latin1')]:
                                        try:
                                            f.seek(0)
                                            df_temp = pd.read_csv(f, sep=sep, encoding=enc)
                                            if len(df_temp.columns) > 1: df_c = df_temp; break
                                        except: pass
                                    if df_c.empty: f.seek(0); df_c = pd.read_csv(f)
                                    cols_str = [str(c).upper() for c in df_c.columns]
                                    if not any(any(x in c for x in ['PROD', 'DESC', 'QTD', 'FISICO', 'RECEB']) for c in cols_str) and df_c.shape[1] >= 9:
                                        f.seek(0); df_bruto = pd.read_csv(f, header=None)
                                        df_c = df_bruto.iloc[:, 5:10].copy() if df_bruto.shape[1] >= 10 else df_bruto.iloc[:, 4:9].copy()
                                        df_c.columns = ['LOJA', 'FORN', 'PROD', 'QTD', 'PADR']
                                else:
                                    todas_abas = pd.read_excel(f, sheet_name=None)
                                    df_c = todas_abas['Contagens'] if 'Contagens' in todas_abas else pd.concat([df for nome, df in todas_abas.items() if 'CARGA' not in str(nome).upper()], ignore_index=True)

                                cols = [str(c).upper().strip() for c in df_c.columns]
                                df_c.columns = cols
                                col_loja = next((c for c in cols if 'LOJA' in c), None)
                                col_prod = next((c for c in cols if 'PROD' in c or 'DESC' in c), None)
                                col_qtd = next((c for c in cols if 'QTD' in c or 'FÍSICO' in c or 'FISICO' in c or 'RECEB' in c), None)
                                col_pad = next((c for c in cols if 'PADR' in c), None)
                                
                                loja_fallback = next((l.capitalize() for l in ['LOJA_1', 'LOJA_2', 'LOJA_3', 'LOJA_5', 'LOJA_6', 'LOJA_7', 'LOJA_8'] if l in f.name.upper()), "Loja_Desconhecida")

                                if col_prod and col_qtd:
                                    for _, row in df_c.iterrows():
                                        loja_str = str(row[col_loja]).upper() if col_loja else loja_fallback.upper()
                                        loja_linha = next((f"Loja_{num}" for num in ['1','2','3','5','6','7','8'] if num in loja_str), loja_fallback)
                                        prod = normalizar(row[col_prod])
                                        qtd = pd.to_numeric(row[col_qtd], errors='coerce')
                                        if prod: contagens_lista.append({'Loja': loja_linha, 'Produto': prod, 'Qtd_Fisico': float(qtd if pd.notna(qtd) else 0.0), 'Padrao_Fisico': str(row[col_pad]) if col_pad and pd.notna(row[col_pad]) else ""})
                            except: pass
                                
                    df_contagens = pd.DataFrame(contagens_lista)
                    if not df_contagens.empty:
                        # --- ESCUDO ANTI-DUPLICIDADE CEGO A FORNECEDOR (Baseado apenas na Loja e no Produto) ---
                        tamanho_original = len(df_contagens)
                        df_contagens = df_contagens.drop_duplicates(subset=['Loja', 'Produto'], keep='last')
                        linhas_removidas = tamanho_original - len(df_contagens)
                        st.success(f"📦 SUCESSO: O motor processou {len(df_contagens)} linhas da Doca! (♻️ {linhas_removidas} duplicidades removidas)")

                    registros = []
                    for (loja, forn_macro), df_ped_group in df_pedidos.groupby(['Loja', 'Fornecedor_Macro']):
                        notas_forn = df_notas_agg[(df_notas_agg['Loja'] == loja) & (df_notas_agg['Fornecedor_Macro'] == forn_macro)] if not df_notas_agg.empty else pd.DataFrame()
                        if notas_forn.empty:
                            for _, ped in df_ped_group.iterrows():
                                stat_v, stat_c, dif = classificar(ped['Qtd'], 0, "SEM_FORNECEDOR")
                                registros.append((id_execucao, loja, ped['Fornecedor_Original'], ped['Produto'], "❌ NOTA NÃO ENCONTRADA", ped['Qtd'], 0, dif, stat_v, stat_c, "-", "-", "⚪ SEM CONTAGEM", 0.0))
                            continue

                        matched_ped_idx, matched_xml_idx, pairs = set(), set(), []
                        for idx_ped, ped in df_ped_group.iterrows():
                            fam_ped = descobrir_familia(ped['Produto'])
                            fam_ampla_ped = fam_ped.split('_')[0] if fam_ped else ""
                            for idx_xml, nota in notas_forn.iterrows():
                                fam_xml = descobrir_familia(nota['Produto'])
                                fam_ampla_xml = fam_xml.split('_')[0] if fam_xml else ""
                                for rw in ["MELANCIA", "BATATA", "CEBOLA", "ALHO"]:
                                    if rw in fam_ampla_ped or rw in fam_ampla_xml:
                                        fam_ampla_ped, fam_ampla_xml = fam_ped, fam_xml
                                        break
                                if fam_ped == fam_xml or fam_ampla_ped == fam_ampla_xml: pairs.append((fuzz.token_sort_ratio(ped['Produto'], nota['Produto']), idx_ped, idx_xml, nota['Produto'], ped['Qtd'], nota['Qtd']))
                        pairs.sort(key=lambda x: x[0], reverse=True)
                        
                        for score, idx_ped, idx_xml, prod_xml, qtd_ped, qtd_fat in pairs:
                            if idx_ped not in matched_ped_idx and idx_xml not in matched_xml_idx:
                                matched_ped_idx.add(idx_ped); matched_xml_idx.add(idx_xml)
                                ped = df_ped_group.loc[idx_ped]
                                stat_v, stat_c, dif = classificar(qtd_ped, qtd_fat, "OK")
                                qtd_fisico, pad_fisico, stat_doca, dif_doca = "-", "-", "⚪ SEM CONTAGEM", 0.0
                                if not df_contagens.empty:
                                    match_fis = df_contagens[(df_contagens['Loja'] == loja) & (df_contagens['Produto'] == normalizar(ped['Produto']))]
                                    if not match_fis.empty:
                                        qtd_fisico = match_fis['Qtd_Fisico'].sum()
                                        pad_fisico = " | ".join([p for p in match_fis['Padrao_Fisico'].unique() if str(p).strip()])
                                        dif_doca = qtd_fisico - qtd_fat
                                        stat_doca = classificar_doca(qtd_fat, qtd_fisico)
                                    else: qtd_fisico, stat_doca, dif_doca = 0.0, classificar_doca(qtd_fat, 0.0), 0.0 - qtd_fat
                                registros.append((id_execucao, loja, ped['Fornecedor_Original'], ped['Produto'], prod_xml, qtd_ped, qtd_fat, dif, stat_v, stat_c, qtd_fisico, pad_fisico, stat_doca, dif_doca))

                        for idx_ped, ped in df_ped_group.iterrows():
                            if idx_ped not in matched_ped_idx:
                                stat_v, stat_c, dif = classificar(ped['Qtd'], 0, "SEM_PRODUTO")
                                qtd_fisico, pad_fisico, stat_doca, dif_doca = "-", "-", "⚪ SEM CONTAGEM", 0.0
                                if not df_contagens.empty:
                                    match_fis = df_contagens[(df_contagens['Loja'] == loja) & (df_contagens['Produto'] == normalizar(ped['Produto']))]
                                    if not match_fis.empty:
                                        qtd_fisico = match_fis['Qtd_Fisico'].sum()
                                        pad_fisico = " | ".join([p for p in match_fis['Padrao_Fisico'].unique() if str(p).strip()])
                                        dif_doca, stat_doca = qtd_fisico - 0, classificar_doca(0, qtd_fisico)
                                    else: qtd_fisico, stat_doca, dif_doca = 0.0, classificar_doca(0, 0.0), 0.0
                                registros.append((id_execucao, loja, ped['Fornecedor_Original'], ped['Produto'], "❌ PRODUTO NÃO FATURADO", ped['Qtd'], 0, dif, stat_v, stat_c, qtd_fisico, pad_fisico, stat_doca, dif_doca))
                        
                        for idx_xml, nota in notas_forn.iterrows():
                            if idx_xml not in matched_xml_idx:
                                prod_xml, qtd_fat = nota['Produto'], nota['Qtd']
                                stat_v, stat_c, dif = f"🟣 SEM PEDIDO (SOBRA {qtd_fat:.2f})".replace('.00',''), 1, qtd_fat
                                qtd_fisico, pad_fisico, stat_doca, dif_doca = "-", "-", "⚪ SEM CONTAGEM", 0.0
                                if not df_contagens.empty: qtd_fisico, stat_doca, dif_doca = 0.0, classificar_doca(qtd_fat, 0.0), 0.0 - qtd_fat
                                registros.append((id_execucao, loja, f"⚠️ {forn_macro} - FATURADO SEM PEDIDO", "❌ NÃO SOLICITADO", prod_xml, 0, qtd_fat, dif, stat_v, stat_c, qtd_fisico, pad_fisico, stat_doca, dif_doca))

                    if registros:
                        df_final = pd.DataFrame(registros, columns=['id_execucao','loja','fornecedor','produto_pedido','produto_xml','qtd_pedido','qtd_nota','diferenca','status_visual','status_codigo','qtd_fisico', 'padrao_fisico', 'status_doca', 'diferenca_doca'])
                        df_final.sort_values(by=['loja', 'fornecedor', 'produto_pedido'], inplace=True)
                        with sqlite3.connect(DB_NAME) as conn: df_final.to_sql("auditoria_v2", conn, if_exists="append", index=False)
                        wb_audit = gerar_excel_auditoria(df_final)
                        out_audit = io.BytesIO()
                        wb_audit.save(out_audit)
                        st.download_button(label="📥 Baixar Auditoria Definitiva (3 Vias)", data=out_audit.getvalue(), file_name=f"Auditoria_3Vias_{id_execucao}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    else: st.error("❌ Erro. Nenhum dado foi processado.")
                except Exception as e: st.error(f"❌ Erro crítico: {e}")
